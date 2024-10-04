import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
import random

# Load the workbook and active sheet
wb = load_workbook('herbalmedicine_1.xlsx', data_only=True, read_only=True)
ws = wb.active

class QuizApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Herbal Medicine Quiz")
        self.root.geometry("600x400")
        self.root.configure(bg="#f0f0f0")  # Set a light background color
        
        # GUI elements with enhanced fonts and padding
        self.question_label = tk.Label(root, text="", font=("Arial", 18, "bold"), bg="#f0f0f0", fg="#333333")
        self.question_label.pack(pady=30)
        
        self.var = tk.IntVar()  # Variable to store the selected radio button value
        self.radio_buttons = []
        
        # Create 5 radio buttons for multiple choice options with larger font and padding
        for i in range(5):
            rb = tk.Radiobutton(root, text="", variable=self.var, value=i, font=("Arial", 14), bg="#f0f0f0", fg="#333333")
            rb.pack(anchor='w', padx=20, pady=5)
            self.radio_buttons.append(rb)
        
        # Styled Submit Button
        self.submit_button = tk.Button(root, text="Submit", command=self.check_answer, font=("Arial", 14, "bold"), bg="#4CAF50", fg="white", padx=20, pady=10)
        self.submit_button.pack(pady=20)
        
        # Styled Next Question Button
        self.next_button = tk.Button(root, text="Next Question", command=self.next_question, font=("Arial", 14, "bold"), bg="#2196F3", fg="white", padx=20, pady=10)
        self.next_button.pack(pady=10)
        
        # Initialize variables
        self.correct_answer_index = None
        self.correct_num = None
        
        # Disable "Next Question" button initially
        self.next_button.config(state=tk.DISABLED)
        
        # Start with the first question
        self.next_question()
    
    def next_question(self):
        # Enable the "Submit" button and disable the "Next Question" button
        self.submit_button.config(state=tk.NORMAL)
        self.next_button.config(state=tk.DISABLED)
        
        # Reset the radio button selection
        self.var.set(-1)
        
        # Pick a random row number between 2 and 177
        num = random.randint(2, 177)
        quiz_question = ws.cell(row=num, column=4).value
        
        if quiz_question:
            self.question_label.config(text=f"생약명이 {quiz_question}인 약용식물의 라틴명은?")
            
            # Get the correct answer
            answer_1 = ws.cell(row=num, column=2).value
            list_x = [num]
            
            # Generate 4 random incorrect answers
            while len(list_x) < 5:
                answer_not_1 = random.randint(2, 177)
                if ws.cell(row=answer_not_1, column=2).value != answer_1:
                    list_x.append(answer_not_1)
            
            # Shuffle the options and save the correct answer index
            random.shuffle(list_x)
            self.correct_answer_index = list_x.index(num)
            self.correct_num = num
            
            # Update the radio buttons with options
            for i, val in enumerate(list_x):
                option = ws.cell(row=val, column=2).value
                self.radio_buttons[i].config(text=option)
        else:
            messagebox.showinfo("Info", "No question available, generating a new one.")
            self.next_question()
    
    def check_answer(self):
        selected = self.var.get()
        if selected == -1:  # Check if no answer was selected
            messagebox.showwarning("Warning", "Please select an answer before submitting.")
            return
        
        if selected == self.correct_answer_index:
            messagebox.showinfo("Result", "정답!")
        else:
            correct_option = ws.cell(row=self.correct_num, column=2).value
            messagebox.showinfo("Result", f"오답! 정답은 {correct_option} 입니다.")
        
        # Disable the "Submit" button after answering, and enable the "Next Question" button
        self.submit_button.config(state=tk.DISABLED)
        self.next_button.config(state=tk.NORMAL)

# Create the main window
root = tk.Tk()
app = QuizApp(root)
root.mainloop()
